"""
AMIC contact search — FastAPI app.

A single password-gated search page over the Supabase `hybrid_search_contacts` RPC.
Query embeddings are generated server-side with bge-base (fastembed); the database
connection lives only on the server, so no DB / service credentials touch the browser.

Env vars (set these in Railway):
  DATABASE_URL   Supabase *session pooler* connection string (IPv4-friendly)
  APP_PASSWORD   the shared password your client types to enter
  SECRET_KEY     random string used to sign the login cookie
  PORT           provided by Railway automatically
"""
import csv
import io
import os
import re
import threading
import uuid

from fastapi import FastAPI, Form, Request
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse, JSONResponse, Response
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.middleware.sessions import SessionMiddleware

from openpyxl import Workbook
from fastembed import TextEmbedding
from psycopg_pool import ConnectionPool

from emailfinder import find_email, domain_from_website

MODEL_NAME = "BAAI/bge-base-en-v1.5"
HERE = os.path.dirname(os.path.abspath(__file__))

DATABASE_URL = os.environ.get("DATABASE_URL", "")
APP_PASSWORD = os.environ.get("APP_PASSWORD", "")
SECRET_KEY = os.environ.get("SECRET_KEY", "change-me-in-railway")

app = FastAPI(title="AMIC Contact Search")
app.add_middleware(SessionMiddleware, secret_key=SECRET_KEY, max_age=60 * 60 * 12)
app.mount("/static", StaticFiles(directory=os.path.join(HERE, "static")), name="static")
templates = Jinja2Templates(directory=os.path.join(HERE, "templates"))

# Loaded once at startup (model + a small connection pool).
_model: TextEmbedding | None = None
_pool: ConnectionPool | None = None
_model_lock = threading.Lock()  # fastembed model is shared across worker threads


@app.on_event("startup")
def _startup():
    global _model, _pool
    _model = TextEmbedding(model_name=MODEL_NAME)
    _pool = ConnectionPool(DATABASE_URL, min_size=1, max_size=4, open=True)


def _authed(request: Request) -> bool:
    return bool(request.session.get("authed"))


def _embed_query(q: str) -> str | None:
    """bge query embedding as a pgvector literal, or None for filter-only browse."""
    q = (q or "").strip()
    if not q:
        return None
    vec = list(_model.query_embed([q]))[0].tolist()
    return "[" + ",".join(f"{x:.6f}" for x in vec) + "]"


def _embed_passages(cards: list[str]) -> list[str]:
    """Embed contact 'cards' as bge passages -> pgvector literals (plain, no prefix)."""
    if not cards:
        return []
    with _model_lock:
        vecs = list(_model.embed(cards, batch_size=64))
    return ["[" + ",".join(f"{x:.6f}" for x in v.tolist()) + "]" for v in vecs]


def _card(full_name, title, organization, tags, notes) -> str:
    """Same card string embed.py uses: full_name | title | organization | tags | notes."""
    tags_joined = " ".join(tags) if tags else ""
    parts = [full_name or "", title or "", organization or "", tags_joined, notes or ""]
    return " | ".join(parts)


def _norm(v):
    v = (v or "").strip() if isinstance(v, str) else v
    return v or None


# ---------------------------------------------------------------- background jobs
# Long operations (embedding an import, SMTP-probing for emails) run in a worker
# thread so the request returns immediately; the client polls /api/jobs/{id}.
_JOBS: dict[str, dict] = {}
_JOBS_LOCK = threading.Lock()


def _job_set(jid: str, **kw):
    with _JOBS_LOCK:
        if jid in _JOBS:
            _JOBS[jid].update(kw)


def _start_job(kind: str, fn) -> str:
    """Spawn fn(progress) in a daemon thread; return a job id to poll."""
    jid = uuid.uuid4().hex[:12]
    with _JOBS_LOCK:
        # keep the store from growing forever across a long-lived process
        if len(_JOBS) > 200:
            for old in [k for k, v in _JOBS.items() if v["status"] != "running"][:100]:
                _JOBS.pop(old, None)
        _JOBS[jid] = {"id": jid, "kind": kind, "status": "running",
                      "done": 0, "total": 0, "message": "Starting…",
                      "result": None, "error": None}

    def progress(**kw):
        _job_set(jid, **kw)

    def runner():
        try:
            result = fn(progress)
            _job_set(jid, status="done", result=result, message="Done", done=_JOBS[jid]["total"])
        except Exception as e:  # surface the failure to the polling client
            _job_set(jid, status="error", error=str(e), message=f"Failed: {e}")

    threading.Thread(target=runner, daemon=True).start()
    return jid


POC_CHOICES = ["TH", "KM", "BW", "DS"]  # Tracy Henke, Kory Mathews, Brandon Wegge, Doug Stuart


def _run_search(q, tags, state, email_status, source, poc, include_dnc,
                semantic_weight, limit):
    params = {
        "qe": _embed_query(q),
        "q": _norm(q),
        "tags": tags or None,
        "state": _norm(state),
        "es": _norm(email_status),
        "src": _norm(source),
        "poc": _norm(poc),
        "dnc": bool(include_dnc),
        "sw": float(semantic_weight),
        "lim": int(limit),
    }
    sql = """
        select contact_id, last_name, first_name, full_name, organization, title,
               email, email_status, phone, linkedin, city, state, tags, source_lists,
               poc, do_not_contact, match_field
        from hybrid_search_contacts(
            query_embedding     => %(qe)s::vector,
            query_text          => %(q)s::text,
            filter_tags         => %(tags)s::text[],
            filter_state        => %(state)s::text,
            filter_email_status => %(es)s::text,
            filter_source       => %(src)s::text,
            filter_poc          => %(poc)s::text,
            include_dnc         => %(dnc)s::boolean,
            semantic_weight     => %(sw)s::real,
            match_count         => %(lim)s::int)
    """
    cols = ["contact_id", "last_name", "first_name", "full_name", "organization",
            "title", "email", "email_status", "phone", "linkedin", "city", "state",
            "tags", "source_lists", "poc", "do_not_contact", "match_field"]
    with _pool.connection() as conn, conn.cursor() as cur:
        cur.execute(sql, params)
        rows = cur.fetchall()
    return [dict(zip(cols, r)) for r in rows]


def _search_args(body):
    """Shared parameter parsing for /api/search and /api/export.csv."""
    return dict(
        q=body.get("q", ""),
        tags=body.get("tags") or None,
        state=body.get("state"),
        email_status=body.get("email_status"),
        source=body.get("source"),
        poc=body.get("poc"),
        include_dnc=bool(body.get("include_dnc")),
        semantic_weight=body.get("semantic_weight", 0.7),
    )


# ---------------------------------------------------------------- auth / pages
@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request, error: str = ""):
    if _authed(request):
        return RedirectResponse("/", status_code=303)
    return templates.TemplateResponse("login.html", {"request": request, "error": error})


@app.post("/login")
def login_submit(request: Request, password: str = Form("")):
    if APP_PASSWORD and password == APP_PASSWORD:
        request.session["authed"] = True
        return RedirectResponse("/", status_code=303)
    return RedirectResponse("/login?error=Incorrect+password", status_code=303)


@app.get("/logout")
def logout(request: Request):
    request.session.clear()
    return RedirectResponse("/login", status_code=303)


@app.get("/", response_class=HTMLResponse)
def home(request: Request):
    if not _authed(request):
        return RedirectResponse("/login", status_code=303)
    return templates.TemplateResponse("search.html", {"request": request})


@app.get("/import", response_class=HTMLResponse)
def import_page(request: Request):
    if not _authed(request):
        return RedirectResponse("/login", status_code=303)
    return templates.TemplateResponse("import.html", {"request": request})


@app.get("/email-finder", response_class=HTMLResponse)
def email_finder_page(request: Request):
    if not _authed(request):
        return RedirectResponse("/login", status_code=303)
    return templates.TemplateResponse("email_finder.html", {"request": request})


@app.get("/healthz")
def healthz():
    return {"ok": True}


@app.get("/api/jobs/{jid}")
def api_job(request: Request, jid: str):
    if not _authed(request):
        return JSONResponse({"error": "unauthorized"}, status_code=401)
    with _JOBS_LOCK:
        job = _JOBS.get(jid)
        if not job:
            return JSONResponse({"error": "unknown job"}, status_code=404)
        return dict(job)


# ---------------------------------------------------------------- api
@app.get("/api/meta")
def api_meta(request: Request):
    if not _authed(request):
        return JSONResponse({"error": "unauthorized"}, status_code=401)
    with _pool.connection() as conn, conn.cursor() as cur:
        cur.execute("select distinct state from contacts where state is not null order by 1")
        states = [r[0] for r in cur.fetchall()]
        cur.execute("""select t, count(*) c from contacts, unnest(tags) t
                       group by t order by c desc, t""")
        tags = [r[0] for r in cur.fetchall()]
        cur.execute("""select s, count(*) c from contacts, unnest(source_lists) s
                       group by s order by c desc, s""")
        sources = [r[0] for r in cur.fetchall()]
    return {"states": states, "tags": tags, "sources": sources,
            "email_statuses": ["valid format", "missing", "check"],
            "poc_choices": POC_CHOICES}


@app.post("/api/search")
async def api_search(request: Request):
    if not _authed(request):
        return JSONResponse({"error": "unauthorized"}, status_code=401)
    body = await request.json()
    # No result cap — return every match ("Max results" control removed from UI).
    rows = _run_search(**_search_args(body),
                       limit=int(body.get("limit") or 1_000_000))
    return {"count": len(rows), "results": rows}


@app.post("/api/poc")
async def api_set_poc(request: Request):
    """Assign / clear a contact's POC. Editable inline from the results table."""
    if not _authed(request):
        return JSONResponse({"error": "unauthorized"}, status_code=401)
    body = await request.json()
    contact_id = _norm(body.get("contact_id"))
    poc = _norm(body.get("poc"))
    if not contact_id:
        return JSONResponse({"error": "missing contact_id"}, status_code=400)
    if poc is not None and poc not in POC_CHOICES:
        return JSONResponse({"error": "invalid poc"}, status_code=400)
    with _pool.connection() as conn, conn.cursor() as cur:
        cur.execute("update contacts set poc = %s where contact_id = %s",
                    (poc, contact_id))
        updated = cur.rowcount
    if not updated:
        return JSONResponse({"error": "unknown contact_id"}, status_code=404)
    return {"ok": True, "contact_id": contact_id, "poc": poc}


@app.post("/api/export.csv")
async def api_export(request: Request):
    if not _authed(request):
        return JSONResponse({"error": "unauthorized"}, status_code=401)
    body = await request.json()
    rows = _export_rows(body)
    fields = ["contact_id", "last_name", "first_name", "full_name", "organization",
              "title", "email", "email_status", "phone", "linkedin", "city", "state",
              "tags", "source_lists", "poc", "do_not_contact"]
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(fields)
    for r in rows:
        row = dict(r)
        row["tags"] = "; ".join(row.get("tags") or [])
        row["source_lists"] = "; ".join(row.get("source_lists") or [])
        row["do_not_contact"] = "yes" if row.get("do_not_contact") else ""
        w.writerow(["" if row.get(f) is None else row.get(f) for f in fields])
    # Prepend a UTF-8 BOM so Excel renders accents/emoji correctly.
    data = "﻿" + buf.getvalue()
    return StreamingResponse(
        iter([data]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=amic_contacts.csv"},
    )


# ---------------------------------------------------------------- saved lists
# Member rows mirror the search result shape so the same grid renders them.
_MEMBER_SQL = """
    select c.contact_id, c.last_name, c.first_name, c.organization, c.title,
           c.email, c.email_status, c.phone, c.linkedin, c.city, c.state,
           c.tags, c.source_lists, c.poc, c.do_not_contact, null::text as match_field
    from contact_list_members m
    join contacts c on c.contact_id = m.contact_id
    where m.list_id = %s
    order by c.last_name asc nulls last, c.full_name asc
"""
_MEMBER_COLS = ["contact_id", "last_name", "first_name", "organization", "title",
                "email", "email_status", "phone", "linkedin", "city", "state",
                "tags", "source_lists", "poc", "do_not_contact", "match_field"]


def _ids(body):
    ids = body.get("contact_ids") or []
    return [str(x) for x in ids if x]


@app.get("/api/lists")
def lists_all(request: Request):
    if not _authed(request):
        return JSONResponse({"error": "unauthorized"}, status_code=401)
    with _pool.connection() as conn, conn.cursor() as cur:
        cur.execute("""select l.id, l.name, count(m.contact_id) as n
                       from contact_lists l
                       left join contact_list_members m on m.list_id = l.id
                       group by l.id, l.name order by lower(l.name)""")
        rows = [{"id": r[0], "name": r[1], "count": r[2]} for r in cur.fetchall()]
    return {"lists": rows}


@app.post("/api/lists")
async def lists_create(request: Request):
    if not _authed(request):
        return JSONResponse({"error": "unauthorized"}, status_code=401)
    body = await request.json()
    name = _norm(body.get("name"))
    if not name:
        return JSONResponse({"error": "missing name"}, status_code=400)
    ids = _ids(body)
    with _pool.connection() as conn, conn.cursor() as cur:
        cur.execute("insert into contact_lists(name) values (%s) returning id", (name,))
        list_id = cur.fetchone()[0]
        if ids:
            cur.execute("""insert into contact_list_members(list_id, contact_id)
                           select %s, x from unnest(%s::text[]) x
                           on conflict do nothing""", (list_id, ids))
    return {"ok": True, "id": list_id, "name": name, "added": len(ids)}


@app.get("/api/lists/{list_id}")
def lists_detail(request: Request, list_id: int):
    if not _authed(request):
        return JSONResponse({"error": "unauthorized"}, status_code=401)
    with _pool.connection() as conn, conn.cursor() as cur:
        cur.execute("select name from contact_lists where id = %s", (list_id,))
        row = cur.fetchone()
        if not row:
            return JSONResponse({"error": "not found"}, status_code=404)
        name = row[0]
        cur.execute(_MEMBER_SQL, (list_id,))
        members = [dict(zip(_MEMBER_COLS, r)) for r in cur.fetchall()]
    return {"id": list_id, "name": name, "count": len(members), "members": members}


@app.post("/api/lists/{list_id}/add")
async def lists_add(request: Request, list_id: int):
    if not _authed(request):
        return JSONResponse({"error": "unauthorized"}, status_code=401)
    ids = _ids(await request.json())
    if not ids:
        return JSONResponse({"error": "no contacts"}, status_code=400)
    with _pool.connection() as conn, conn.cursor() as cur:
        cur.execute("select 1 from contact_lists where id = %s", (list_id,))
        if not cur.fetchone():
            return JSONResponse({"error": "not found"}, status_code=404)
        cur.execute("""insert into contact_list_members(list_id, contact_id)
                       select %s, x from unnest(%s::text[]) x
                       on conflict do nothing""", (list_id, ids))
        added = cur.rowcount
    return {"ok": True, "added": added}


@app.post("/api/lists/{list_id}/remove")
async def lists_remove(request: Request, list_id: int):
    if not _authed(request):
        return JSONResponse({"error": "unauthorized"}, status_code=401)
    ids = _ids(await request.json())
    if not ids:
        return JSONResponse({"error": "no contacts"}, status_code=400)
    with _pool.connection() as conn, conn.cursor() as cur:
        cur.execute("delete from contact_list_members where list_id = %s and contact_id = any(%s::text[])",
                    (list_id, ids))
        removed = cur.rowcount
    return {"ok": True, "removed": removed}


@app.post("/api/lists/{list_id}/rename")
async def lists_rename(request: Request, list_id: int):
    if not _authed(request):
        return JSONResponse({"error": "unauthorized"}, status_code=401)
    name = _norm((await request.json()).get("name"))
    if not name:
        return JSONResponse({"error": "missing name"}, status_code=400)
    with _pool.connection() as conn, conn.cursor() as cur:
        cur.execute("update contact_lists set name = %s where id = %s", (name, list_id))
        if not cur.rowcount:
            return JSONResponse({"error": "not found"}, status_code=404)
    return {"ok": True, "name": name}


@app.post("/api/lists/{list_id}/delete")
def lists_delete(request: Request, list_id: int):
    if not _authed(request):
        return JSONResponse({"error": "unauthorized"}, status_code=401)
    with _pool.connection() as conn, conn.cursor() as cur:
        cur.execute("delete from contact_lists where id = %s", (list_id,))
        if not cur.rowcount:
            return JSONResponse({"error": "not found"}, status_code=404)
    return {"ok": True}


# ================================================================ bulk import
# The browser parses the CSV/XLSX and POSTs already-mapped rows as JSON, so the
# server only validates, dedupes, inserts, and embeds. Importable fields:
IMPORT_FIELDS = ["full_name", "first_name", "last_name", "title", "organization",
                 "email", "email_status", "phone", "linkedin", "city", "state",
                 "tags", "source_lists", "notes"]
_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


def _nk(s) -> str:
    """Normalized key for dedupe: lowercased, whitespace-collapsed."""
    return re.sub(r"\s+", " ", (s or "").strip().lower())


def _as_list(v):
    """Coerce a tags/source cell (list, or ';'/',' delimited string) to a clean list."""
    if v is None:
        return []
    if isinstance(v, list):
        return [str(x).strip() for x in v if str(x).strip()]
    return [p.strip() for p in re.split(r"[;,]", str(v)) if p.strip()]


def _norm_import_row(raw: dict) -> dict:
    """One incoming record -> normalized dict over IMPORT_FIELDS."""
    r = {k: (_norm(raw.get(k)) if k not in ("tags", "source_lists") else _as_list(raw.get(k)))
         for k in IMPORT_FIELDS}
    if not r["full_name"]:
        r["full_name"] = " ".join(x for x in [r["first_name"], r["last_name"]] if x) or None
    return r


def _load_dedupe_maps(cur):
    """email(lower) -> contact_id, and (name|org)(lower) -> contact_id."""
    cur.execute("select contact_id, email, full_name, organization from contacts")
    by_email, by_nameorg = {}, {}
    for cid, email, fn, org in cur.fetchall():
        if email and email.strip():
            by_email[email.strip().lower()] = cid
        by_nameorg[_nk(fn) + "|" + _nk(org)] = cid
    return by_email, by_nameorg


def _classify(row: dict, by_email: dict, by_nameorg: dict):
    """('new'|'duplicate', matched_on, existing_contact_id)."""
    em = (row.get("email") or "").strip().lower()
    if em and em in by_email:
        return "duplicate", "email", by_email[em]
    key = _nk(row.get("full_name")) + "|" + _nk(row.get("organization"))
    if key in by_nameorg:
        return "duplicate", "name + org", by_nameorg[key]
    return "new", None, None


# ---- smart parser for pasted "asinine" lists -----------------------------
# Turns raw text copied straight out of an email into IMPORT_FIELDS rows that
# flow through the same preview/commit/embed pipeline as an uploaded file.
_EMAIL_FIND = re.compile(r"[^@\s<>;,()]+@[^@\s<>;,()]+\.[^@\s<>;,()]+")
_LINE_SEP = re.compile(r"\s+[–—|]\s+|\s+-\s+|\t+")
_GENERIC_DOMAINS = {"gmail.com", "yahoo.com", "outlook.com", "hotmail.com",
                    "aol.com", "icloud.com", "me.com", "msn.com", "live.com",
                    "comcast.net", "sbcglobal.net", "att.net", "verizon.net"}


def _split_name(name: str):
    """'Last, First' or 'First Last' -> (first, last)."""
    name = re.sub(r"\s+", " ", (name or "").strip()).strip('"').strip()
    if not name:
        return None, None
    if "," in name:
        last, _, first = name.partition(",")
        return (first.strip() or None), (last.strip() or None)
    parts = name.split(" ")
    if len(parts) == 1:
        return parts[0], None
    return parts[0], " ".join(parts[1:])


def _org_from_domain(email: str | None):
    """Honest placeholder org from the email domain (user can rename in review)."""
    if not email or "@" not in email:
        return None
    dom = email.split("@", 1)[1].lower().strip()
    if not dom or dom in _GENERIC_DOMAINS:
        return None
    return dom


def _parse_pasted(text: str, derive_org: bool = False) -> list[dict]:
    text = text or ""
    rows: list[dict] = []
    if _EMAIL_FIND.search(text):
        # Address-blob mode: Outlook-style "Name <email>; …" or bare emails,
        # separated by ';' or newlines.
        for chunk in re.split(r"[;\n]+", text):
            chunk = chunk.strip().strip(",").strip()
            if not chunk:
                continue
            m = re.match(r'^(.*?)<\s*([^<>\s]+@[^<>\s]+)\s*>\s*$', chunk)
            if m:
                name, email = m.group(1), m.group(2)
            else:
                em = _EMAIL_FIND.search(chunk)
                if not em:
                    name, email = chunk, None
                elif em.group(0) == chunk:
                    name, email = "", chunk
                else:                       # "Name email" with no brackets
                    email = em.group(0)
                    name = chunk.replace(email, "").strip(' <>"\t')
            first, last = _split_name(name) if name else (None, None)
            full = " ".join(x for x in [first, last] if x) or (email or None)
            if not full:
                continue
            rows.append({"full_name": full, "first_name": first, "last_name": last,
                         "email": email,
                         "organization": _org_from_domain(email) if derive_org else None})
    else:
        # Line-list mode: one entry per line, "Name – Organization".
        for line in text.splitlines():
            line = line.strip()
            if not line:
                continue
            parts = _LINE_SEP.split(line, maxsplit=1)
            name = parts[0].strip()
            org = parts[1].strip() if len(parts) > 1 else None
            first, last = _split_name(name)
            full = " ".join(x for x in [first, last] if x) or name or None
            if not full:
                continue
            rows.append({"full_name": full, "first_name": first, "last_name": last,
                         "organization": org})
    return rows


@app.post("/api/import/parse")
async def api_import_parse(request: Request):
    """Structure raw pasted text into importable rows (no DB writes)."""
    if not _authed(request):
        return JSONResponse({"error": "unauthorized"}, status_code=401)
    body = await request.json()
    rows = _parse_pasted(body.get("text") or "", bool(body.get("derive_org")))
    mode = "address-blob" if any(r.get("email") for r in rows) else "name-org"
    return {"rows": rows, "count": len(rows), "mode": mode}


@app.post("/api/import/preview")
async def api_import_preview(request: Request):
    """Classify incoming rows as new vs. likely-duplicate without writing anything."""
    if not _authed(request):
        return JSONResponse({"error": "unauthorized"}, status_code=401)
    rows = (await request.json()).get("rows") or []
    norm = [_norm_import_row(r) for r in rows]
    with _pool.connection() as conn, conn.cursor() as cur:
        by_email, by_nameorg = _load_dedupe_maps(cur)
    out, n_new, n_dupe, n_err = [], 0, 0, 0
    seen_in_file = set()
    for i, r in enumerate(norm):
        if not r["full_name"]:
            out.append({"idx": i, "status": "error", "matched_on": "no name",
                        "full_name": None, "organization": r["organization"],
                        "email": r["email"]})
            n_err += 1
            continue
        status, on, _ = _classify(r, by_email, by_nameorg)
        # also catch duplicates within the uploaded file itself
        fk = _nk(r["full_name"]) + "|" + _nk(r["organization"])
        if status == "new" and fk in seen_in_file:
            status, on = "duplicate", "in this file"
        seen_in_file.add(fk)
        n_new += status == "new"
        n_dupe += status == "duplicate"
        out.append({"idx": i, "status": status, "matched_on": on,
                    "full_name": r["full_name"], "organization": r["organization"],
                    "email": r["email"]})
    return {"total": len(norm), "new": n_new, "duplicates": n_dupe,
            "errors": n_err, "fields": IMPORT_FIELDS, "rows": out}


def _email_status_for(email: str | None) -> str:
    if not email:
        return "missing"
    return "valid format" if _EMAIL_RE.match(email) else "check"


def _commit_job(rows, policy, source_label):
    norm = [_norm_import_row(r) for r in rows if (r.get("full_name") or r.get("first_name")
            or r.get("last_name"))]

    def job(progress):
        inserted = merged = skipped = 0
        affected_ids: list[str] = []
        with _pool.connection() as conn, conn.cursor() as cur:
            by_email, by_nameorg = _load_dedupe_maps(cur)
            cur.execute("select lower(name), org_id from organizations")
            org_ids = {n: o for n, o in cur.fetchall()}
            cur.execute("select coalesce(max((substring(contact_id from 2))::int), 0) "
                        "from contacts where contact_id ~ '^C[0-9]+$'")
            next_n = cur.fetchone()[0] + 1

            progress(total=len(norm), message="Writing contacts…")
            for i, r in enumerate(norm):
                if not r["full_name"]:
                    skipped += 1
                    progress(done=i + 1)
                    continue
                status, _, existing_id = _classify(r, by_email, by_nameorg)
                src = r["source_lists"] + ([source_label] if source_label
                                           and source_label not in r["source_lists"] else [])
                if status == "new" or policy == "all":
                    cid = f"C{next_n:04d}"
                    next_n += 1
                    oid = org_ids.get(_nk(r["organization"]))
                    email = r["email"]
                    cur.execute("""
                        insert into contacts
                          (contact_id, first_name, last_name, full_name, title, organization,
                           org_id, email, email_status, phone, linkedin, city, state,
                           tags, source_lists, email_source, do_not_contact, notes, embedding)
                        values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,false,%s,null)
                    """, (cid, r["first_name"], r["last_name"], r["full_name"], r["title"],
                          r["organization"], oid, email,
                          r["email_status"] or _email_status_for(email), r["phone"],
                          r["linkedin"], r["city"], r["state"], r["tags"] or None,
                          src or None, "import" if email else None, r["notes"]))
                    inserted += 1
                    affected_ids.append(cid)
                    # keep in-batch dedupe maps current so later rows see this insert
                    if email:
                        by_email[email.strip().lower()] = cid
                    by_nameorg[_nk(r["full_name"]) + "|" + _nk(r["organization"])] = cid
                elif policy == "merge":
                    cur.execute("""
                        update contacts set
                          first_name   = coalesce(nullif(first_name,''),   %(first)s),
                          last_name    = coalesce(nullif(last_name,''),     %(last)s),
                          title        = coalesce(nullif(title,''),         %(title)s),
                          organization = coalesce(nullif(organization,''),  %(org)s),
                          email        = coalesce(nullif(email,''),         %(email)s),
                          phone        = coalesce(nullif(phone,''),         %(phone)s),
                          linkedin     = coalesce(nullif(linkedin,''),      %(linkedin)s),
                          city         = coalesce(nullif(city,''),          %(city)s),
                          state        = coalesce(nullif(state,''),         %(state)s),
                          notes        = coalesce(nullif(notes,''),         %(notes)s),
                          tags         = (select array(select distinct e from
                                          unnest(coalesce(tags,'{}'::text[]) || %(tags)s::text[]) e where e <> '')),
                          source_lists = (select array(select distinct e from
                                          unnest(coalesce(source_lists,'{}'::text[]) || %(src)s::text[]) e where e <> ''))
                        where contact_id = %(cid)s
                    """, {"first": r["first_name"], "last": r["last_name"], "title": r["title"],
                          "org": r["organization"], "email": r["email"], "phone": r["phone"],
                          "linkedin": r["linkedin"], "city": r["city"], "state": r["state"],
                          "notes": r["notes"], "tags": r["tags"], "src": src, "cid": existing_id})
                    cur.execute("""insert into merge_log(contact_id, full_name, rows_merged, sources)
                                   values (%s,%s,1,%s)""",
                                (existing_id, r["full_name"], source_label or "bulk import"))
                    merged += 1
                    affected_ids.append(existing_id)
                else:  # policy == "skip"
                    skipped += 1
                progress(done=i + 1)

            # Embed everything we touched that now lacks a vector (new rows + any
            # merged rows whose re-embed trigger nulled the embedding).
            if affected_ids:
                progress(message="Generating embeddings…")
                cur.execute("""select contact_id, full_name, title, organization, tags, notes
                               from contacts where contact_id = any(%s) and embedding is null""",
                            (affected_ids,))
                todo = cur.fetchall()
                if todo:
                    cards = [_card(t[1], t[2], t[3], t[4], t[5]) for t in todo]
                    vecs = _embed_passages(cards)
                    cur.execute("create temp table _imp_emb (contact_id text primary key, "
                                "embedding vector(768)) on commit drop")
                    with cur.copy("copy _imp_emb (contact_id, embedding) from stdin") as cp:
                        for (cid, *_), lit in zip(todo, vecs):
                            cp.write_row((cid, lit))
                    cur.execute("update contacts c set embedding = s.embedding "
                                "from _imp_emb s where c.contact_id = s.contact_id")
            conn.commit()
        return {"inserted": inserted, "merged": merged, "skipped": skipped,
                "embedded": len(affected_ids), "total": len(rows)}

    return job


@app.post("/api/import/commit")
async def api_import_commit(request: Request):
    """Insert/merge the uploaded rows and embed them. Returns a job id to poll."""
    if not _authed(request):
        return JSONResponse({"error": "unauthorized"}, status_code=401)
    body = await request.json()
    rows = body.get("rows") or []
    policy = body.get("policy", "skip")
    if policy not in ("skip", "merge", "all"):
        return JSONResponse({"error": "bad policy"}, status_code=400)
    if not rows:
        return JSONResponse({"error": "no rows"}, status_code=400)
    source_label = _norm(body.get("source_label"))
    jid = _start_job("import", _commit_job(rows, policy, source_label))
    return {"job": jid}


# ================================================================ email finder
@app.post("/api/email/find")
async def api_email_find(request: Request):
    """
    Find emails for contacts missing one. Background job; poll /api/jobs/{id}.
    Body: contact_ids[] (or none -> all missing, capped), and options:
      hunter_key     str  per-request Hunter.io key (overrides the env key)
      local_only     bool skip the API arm, pattern-guess only
      min_confidence float don't write candidates below this score (0..1)
    """
    if not _authed(request):
        return JSONResponse({"error": "unauthorized"}, status_code=401)
    body = await request.json()
    ids = [str(x) for x in (body.get("contact_ids") or []) if x]
    limit = min(int(body.get("limit", 500)), 1500)
    local_only = bool(body.get("local_only"))
    min_conf = float(body.get("min_confidence") or 0)
    # "" forces the local pattern arm; None lets find_email fall back to the env key.
    key = "" if local_only else (_norm(body.get("hunter_key")))

    def job(progress):
        with _pool.connection() as conn, conn.cursor() as cur:
            if ids:
                cur.execute("""
                    select c.contact_id, c.first_name, c.last_name, c.full_name,
                           c.organization, o.website
                    from contacts c left join organizations o on o.org_id = c.org_id
                    where c.contact_id = any(%s)
                      and (c.email is null or c.email = '')""", (ids,))
            else:
                cur.execute("""
                    select c.contact_id, c.first_name, c.last_name, c.full_name,
                           c.organization, o.website
                    from contacts c left join organizations o on o.org_id = c.org_id
                    where (c.email is null or c.email = '')
                    order by c.contact_id limit %s""", (limit,))
            targets = cur.fetchall()

        found = skipped_low = no_match = 0
        details = []
        progress(total=len(targets), message=f"Searching {len(targets)} contacts…")
        for i, (cid, first, last, full, org, website) in enumerate(targets):
            domain = domain_from_website(website)
            hit = None
            try:
                hit = find_email(first, last, full, domain, company=org, hunter_key=key)
            except Exception:
                hit = None
            if hit and hit.confidence >= min_conf:
                with _pool.connection() as conn, conn.cursor() as cur:
                    cur.execute("""update contacts
                                   set email = %s, email_status = 'check',
                                       email_source = %s, email_confidence = %s
                                   where contact_id = %s and (email is null or email = '')""",
                                (hit.email, hit.source, hit.confidence, cid))
                found += 1
                details.append({"contact_id": cid, "full_name": full, "email": hit.email,
                                "confidence": hit.confidence, "source": hit.source,
                                "detail": hit.detail})
            elif hit:
                skipped_low += 1
                details.append({"contact_id": cid, "full_name": full, "email": None,
                                "confidence": hit.confidence, "source": hit.source,
                                "detail": f"below threshold ({hit.confidence:.2f})"})
            else:
                no_match += 1
                details.append({"contact_id": cid, "full_name": full, "email": None,
                                "confidence": 0, "source": None,
                                "detail": "no domain" if not domain else "no match"})
            progress(done=i + 1, message=f"{found} found / {i + 1} checked")
        return {"checked": len(targets), "found": found,
                "skipped_low": skipped_low, "no_match": no_match, "details": details}

    jid = _start_job("email", job)
    return {"job": jid}


@app.post("/api/email/apply")
async def api_email_apply(request: Request):
    """Bulk-confirm or clear found emails. action: 'accept' | 'clear'."""
    if not _authed(request):
        return JSONResponse({"error": "unauthorized"}, status_code=401)
    body = await request.json()
    ids = [str(x) for x in (body.get("contact_ids") or []) if x]
    action = body.get("action")
    if not ids:
        return JSONResponse({"error": "no contacts"}, status_code=400)
    with _pool.connection() as conn, conn.cursor() as cur:
        if action == "accept":   # promote guessed (check) emails to verified
            cur.execute("""update contacts set email_status = 'valid format'
                           where contact_id = any(%s) and email_status = 'check'""", (ids,))
        elif action == "clear":  # discard a bad guess, back to missing
            cur.execute("""update contacts
                           set email = null, email_status = 'missing',
                               email_source = null, email_confidence = null
                           where contact_id = any(%s)""", (ids,))
        else:
            return JSONResponse({"error": "bad action"}, status_code=400)
        n = cur.rowcount
    return {"ok": True, "updated": n}


# ================================================================ exports
def _export_rows(body: dict) -> list[dict]:
    """Rows for an export: explicit contact_ids if given, else the current search."""
    ids = [str(x) for x in (body.get("contact_ids") or []) if x]
    if ids:
        cols = ["contact_id", "last_name", "first_name", "full_name", "organization",
                "title", "email", "email_status", "phone", "linkedin", "city", "state",
                "tags", "source_lists", "poc", "do_not_contact"]
        with _pool.connection() as conn, conn.cursor() as cur:
            cur.execute(f"""select {','.join(cols)} from contacts
                            where contact_id = any(%s)
                            order by last_name asc nulls last, full_name asc""", (ids,))
            rows = [dict(zip(cols, r)) for r in cur.fetchall()]
    else:
        rows = _run_search(**_search_args(body), limit=min(int(body.get("limit", 2000)), 5000))
    # invite-list hygiene: drop no-email rows / collapse duplicate addresses
    if body.get("require_email"):
        rows = [r for r in rows if (r.get("email") or "").strip()]
    if body.get("dedupe_email"):
        seen, out = set(), []
        for r in rows:
            e = (r.get("email") or "").strip().lower()
            if e and e in seen:
                continue
            if e:
                seen.add(e)
            out.append(r)
        rows = out
    return rows


def _greeting(r: dict) -> str:
    return f"Dear {r.get('first_name') or r.get('full_name') or 'Colleague'}"


@app.post("/api/export.xlsx")
async def api_export_xlsx(request: Request):
    """Outlook / Word mail-merge sheet: one row per person, greeting column ready."""
    if not _authed(request):
        return JSONResponse({"error": "unauthorized"}, status_code=401)
    body = await request.json()
    rows = _export_rows(body)
    wb = Workbook()
    ws = wb.active
    ws.title = "Invitees"
    headers = ["First Name", "Last Name", "Greeting", "Email", "Organization",
               "Title", "City", "State", "POC"]
    ws.append(headers)
    for r in rows:
        ws.append([r.get("first_name") or "", r.get("last_name") or "", _greeting(r),
                   r.get("email") or "", r.get("organization") or "", r.get("title") or "",
                   r.get("city") or "", r.get("state") or "", r.get("poc") or ""])
    for col, w in zip("ABCDEFGHI", (16, 16, 22, 30, 30, 26, 16, 8, 8)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=amic_invitees.xlsx"})


@app.post("/api/export/outlook.csv")
async def api_export_outlook(request: Request):
    """CSV with the exact headers Outlook 'Import Contacts' expects."""
    if not _authed(request):
        return JSONResponse({"error": "unauthorized"}, status_code=401)
    body = await request.json()
    rows = _export_rows(body)
    fields = ["First Name", "Last Name", "E-mail Address", "Company", "Job Title",
              "Business Phone", "Business City", "Business State", "Web Page", "Notes"]
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(fields)
    for r in rows:
        w.writerow([r.get("first_name") or "", r.get("last_name") or "", r.get("email") or "",
                    r.get("organization") or "", r.get("title") or "", r.get("phone") or "",
                    r.get("city") or "", r.get("state") or "", r.get("linkedin") or "",
                    "; ".join(r.get("source_lists") or [])])
    data = "﻿" + buf.getvalue()
    return StreamingResponse(iter([data]), media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=amic_outlook_contacts.csv"})


@app.post("/api/export/bcc")
async def api_export_bcc(request: Request):
    """Semicolon-joined address string for a quick small Outlook send."""
    if not _authed(request):
        return JSONResponse({"error": "unauthorized"}, status_code=401)
    body = dict(await request.json())
    body["require_email"] = True
    body["dedupe_email"] = True
    rows = _export_rows(body)
    emails = [r["email"].strip() for r in rows if (r.get("email") or "").strip()]
    return {"count": len(emails), "emails": "; ".join(emails)}


def _sheet_from_query(ws, headers, cur, sql):
    """Write a header row + every result row; arrays become '; '-joined strings."""
    ws.append(headers)
    cur.execute(sql)
    for row in cur.fetchall():
        ws.append(["; ".join(map(str, v)) if isinstance(v, list) else v for v in row])
    ws.freeze_panes = "A2"


@app.get("/api/export/database.xlsx")
def api_export_database(request: Request):
    """Download the entire database as the familiar multi-tab workbook (live snapshot)."""
    if not _authed(request):
        return JSONResponse({"error": "unauthorized"}, status_code=401)
    wb = Workbook()
    with _pool.connection() as conn, conn.cursor() as cur:
        ws = wb.active
        ws.title = "Contacts"
        _sheet_from_query(ws,
            ["contact_id", "last_name", "first_name", "organization", "title",
             "full_name", "org_id", "email", "email_status", "phone", "linkedin",
             "city", "state", "location_raw", "tags", "source_lists", "POC",
             "do_not_contact", "last_contacted", "notes", "needs_review"], cur,
            """select contact_id, last_name, first_name, organization, title, full_name,
                      org_id, email, email_status, phone, linkedin, city, state,
                      location_raw, tags, source_lists, poc, do_not_contact,
                      last_contacted, notes, needs_review
               from contacts order by (substring(contact_id from 2))::int""")
        _sheet_from_query(wb.create_sheet("Organizations"),
            ["org_id", "name", "category", "website", "email", "phone", "address",
             "city", "state", "zip", "employees_local", "employees_companywide",
             "source_lists", "notes"], cur,
            """select org_id, name, category, website, email, phone, address, city,
                      state, zip, employees_local, employees_companywide, source_lists, notes
               from organizations order by org_id""")
        _sheet_from_query(wb.create_sheet("Memberships"),
            ["contact_id", "full_name", "source_file", "list_category",
             "email_as_listed", "org_as_listed"], cur,
            """select m.contact_id, c.full_name, m.source_file, m.list_category,
                      m.email_as_listed, m.org_as_listed
               from memberships m left join contacts c on c.contact_id = m.contact_id
               order by m.contact_id""")
        _sheet_from_query(wb.create_sheet("Merge Log"),
            ["contact_id", "full_name", "rows_merged", "sources", "original_rows"], cur,
            """select contact_id, full_name, rows_merged, sources, original_rows
               from merge_log order by contact_id""")
    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Contact Database.xlsx"})
